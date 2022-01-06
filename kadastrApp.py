import os
import shutil
import sqlite3
import sys
import configparser
from docxtpl import DocxTemplate
import openpyxl
from fuzzywuzzy import fuzz
from os import listdir
from datetime import date
from datetime import datetime
from PyQt5 import QtWidgets
from PyQt5.QtCore import QDate, QRegExp
from PyQt5.QtGui import QColor, QBrush, QTextCharFormat, QRegExpValidator, QIcon
from PyQt5.QtWidgets import QTableWidgetItem, QMessageBox, QInputDialog, QFileDialog

import fileUi.newForm
import fileUi.formAddCity


class startWindow(QtWidgets.QMainWindow, fileUi.newForm.Ui_MainWindow):
    def __init__(self):
        super().__init__()

        # window setting
        self.setupUi(self)
        self.tableWidget.hideColumn(9)
        self.configParser()

        os.chdir(self.configParser()['PATH']['pathToWorkFlow'])
        self.dbName = self.configParser()['PATH']['pathToDateBase']
        self.pathDoc = self.configParser()['PATH']['pathToDocFile']

        # var

        self.infoClientForSelectCalendar = None
        self.infoAddressForSelectCalendar = None
        self.infoWorkForSelectCalendar = None
        self.idClientForSelectcalendar = None
        self.workForCalendar = None
        self.statusForCalendar = None
        self.dateStatusForCalendar = None
        self.createPath = None
        self.infoClient = None
        self.infoPassportClient = None
        self.infoWorkClient = None
        self.infoDocFillInfo = None
        self.indexClient = None
        self.newCity = None
        self.pathFolderNew = None
        self.allInfoClient = None
        self.allInfoAddress = None
        self.allInfoWork = None
        self.allInfoCity = None
        self.allDocFillInfo = None
        self.lastIndex = None
        self.pathFolderClient = None


        # get info in date base

        self.getAllClientInfo()
        self.getAllInfoCity()

        # completer

        # regExp
        self.regExpPhone = QRegExp('^\d\d\d\d\d\d\d\d\d\d\d$')
        self.intValidatorPhone = QRegExpValidator(self.regExpPhone, self.lineEditTelefone)
        self.lineEditTelefone.setValidator(self.intValidatorPhone)
        self.intValidatorPhone = QRegExpValidator(self.regExpPhone, self.lineEditTelefoneAdd)
        self.lineEditTelefoneAdd.setValidator(self.intValidatorPhone)

        # page client

        self.pushButtonDelete.clicked.connect(self.deleteClient)
        self.pushButtonSearch.clicked.connect(self.searchClient)
        self.lineEditSearch.returnPressed.connect(self.pushButtonSearch.click)
        self.pushButtonUpdate.clicked.connect(lambda: self.updateTableClientToId(self.getAllIdClient()))
        self.pushButtonViewReady.clicked.connect(self.clientInfoReady)
        self.calendarWidget.clicked.connect(self.calendarSelectWork)
        self.pushButtonViewForgotten.clicked.connect(self.clientInfoDebts)
        self.tableWidget.clicked.connect(self.calendarSelectClient)

        # page add client

        self.pushButtonChangeCity.clicked.connect(self.updateInfoCity)
        self.pushButtonAdd.clicked.connect(self.addClient)
        self.pushButtonAddCity.clicked.connect(self.insertCity)

        # page update client

        self.pushButtonOpenFolder.clicked.connect(lambda: self.openFolder(self.infoClient[0][5]))
        self.pushButtonEdit.clicked.connect(self.updateFullInfoForClient)
        self.pushButtonChangePath.clicked.connect(self.updatePathToClient)
        self.tableWidget.doubleClicked.connect(self.fillClientInfo)
        self.pushButtonChangeWorkStatus.clicked.connect(self.changeWorkInfo)
        self.checkBoxPrepayment.clicked.connect(self.prepayment)
        self.checkBoxRemains.clicked.connect(self.remains)

        # work client info

        self.pushButtonFillDoc.clicked.connect(self.fillDocClient)



    def getDirectoryToWorkFlow(self):  # <-----
        dirlist = QFileDialog.getExistingDirectory(self,
                                                   "Выберите сетевую папку, где будет хранится информация о клиентах",
                                                   ".")
        return dirlist

    def getDirectoryDoc(self):  # <-----
        dirlist = QFileDialog.getExistingDirectory(self,
                                                   "Выберите сетевую папку, где расположены документы, готовые для автоматического заполнения",
                                                   ".")
        return dirlist

    def getFileDateBase(self):
        filename, filetype = QFileDialog.getOpenFileName(self,
                                                         "Выберите файл базы данных( Необходим сетевой путь)",
                                                         ".",
                                                         "All Files(*)")
        return filename

    def configParser(self):
        if configparser.ConfigParser().read('config.ini') == []:
            config = configparser.ConfigParser()
            config['PATH'] = {'pathToDateBase': os.path.abspath(self.getFileDateBase()),
                              'pathToDocFile': os.path.abspath(self.getDirectoryDoc()),
                              'pathToWorkFlow': os.path.abspath(self.getDirectoryToWorkFlow())}
            with open('config.ini', 'w') as configfile:
                config.write(configfile)
            config.read('config.ini')
            return config
        else:
            config = configparser.ConfigParser()
            config.read('config.ini')
            return config

    def msgInfo(self, text):
        QMessageBox.information(self, "Информация",
                                text)

    def querySelect(self, sql, param=None):
        try:
            con = sqlite3.connect(self.dbName)
            with con:
                if param is None:
                    cur = con.cursor()
                    cur.execute(sql)
                    res = cur.fetchall()


                else:
                    cur = con.cursor()
                    cur.execute(sql, param)
                    res = cur.fetchall()
            if con:
                con.close()
            return res
        except sqlite3.Error as err:
            sqlError = 'Sql error: %s' % (' '.join(err.args))
            errorClass = "Exception class is: ", err.__class__
            self.msgInfo(str(sqlError) + "; " + str(errorClass))

    def queryInsert(self, sql, param=None):
        try:
            con = sqlite3.connect(self.dbName)
            with con:
                if param is None:
                    cur = con.cursor()
                    cur.execute("PRAGMA foreign_keys=ON")
                    cur.execute(sql)

                    con.commit()

                    cur.execute("SELECT last_insert_rowid()")
                    index = cur.fetchone()
                    self.lastIndex = index
                else:
                    cur = con.cursor()
                    cur.execute("PRAGMA foreign_keys=ON")
                    cur.execute(sql, param)
                    con.commit()
                    cur.execute("SELECT last_insert_rowid()")
                    index = cur.fetchone()
                    self.lastIndex = index
            if con:
                con.close()
        except sqlite3.Error as err:
            sqlError = 'Sql error: %s' % (' '.join(err.args))
            errorClass = "Exception class is: ", err.__class__
            self.msgInfo(str(sqlError) + "; " + str(errorClass))

    def createDateObject(self, date):
        dateBuild = QDate()
        year = int(date[6:10])
        mnt = int(date[3:5])
        day = int(date[0:2])
        dateBuild.setDate(year, mnt, day)
        return dateBuild

    #

    def getInfoForSelectCalendar(self, id_client):  # Доделать
        sqlInfo = ""
        sqlAddress = ""
        sqlWork = ""
        self.infoClientForSelectCalendar = self.querySelect(sqlInfo, id_client)
        self.infoAddressForSelectCalendar = self.querySelect(sqlAddress, id_client)
        self.infoWorkForSelectCalendar = self.querySelect(sqlWork, id_client)

    def getIdClientForSelectCalendar(self, date_status):
        sqlDate = "select id_client from work_info_client where date_status = ?"
        result = self.querySelect(sqlDate, (date_status,))
        return result

    def getIdService(self, id):

        sqlIdService = "SELECT service from info_client where id = ?"

        return self.querySelect(sqlIdService, (id,))

    def getIdDocInfoClient(self, id_client):

        sqlIdDocInfo = "SELECT * from doc_info_client where id_client = ?"

        return self.querySelect(sqlIdDocInfo, (id_client,))

    def getIdCity(self, name):

        for i in range(len(self.allInfoCity)):
            if self.allInfoCity[i][1] == name:
                return self.allInfoCity[i][0]

    def getWorkInfoClient(self, id_client):

        sqlIdAddressInfo = "SELECT * from work_info_client where id_client = ?"

        return self.querySelect(sqlIdAddressInfo, (id_client,))

    def getAddressInfoClient(self, id_client):

        sqlIdAddressInfo = "SELECT * from address_info_client where id_client = ?"

        return self.querySelect(sqlIdAddressInfo, (id_client,))

    def getDocInfoClient(self, id_client):

        sqlDocInfo = "SELECT * from doc_info_client where id_client = ?"

        return self.querySelect(sqlDocInfo, (id_client,))

    def getDocFillInfo(self, id_client):

        sqlDocFillInfo = "SELECT * from doc_fill_info where id_client = ?"

        return self.querySelect(sqlDocFillInfo, (id_client,))

    def getInfoReady(self):
        sqlStatus = "select id_client from work_info_client where status = 1"
        resultStatus = self.querySelect(sqlStatus)
        massClientToReady = []
        for item in resultStatus:
            massClientToReady.append(item[0])
        return massClientToReady

    def getNameCity(self, id_city):
        sqlCity = "select city_name from city where id = ?"
        return self.querySelect(sqlCity, (id_city,))

    def getAllClientInfo(self):
        # list of page client

        sqlInfoClient = "SELECT * from info_client"
        self.allInfoClient = self.querySelect(sqlInfoClient)

        sqlInfoAddress = "SELECT * from address_info_client"
        self.allInfoAddress = self.querySelect(sqlInfoAddress)

        sqlInfoWork = "SELECT * from work_info_client"
        self.allInfoWork = self.querySelect(sqlInfoWork)

        sqlFillInfo = "SELECT * from doc_fill_info"
        self.allDocFillInfo = self.querySelect(sqlFillInfo)

    def getAllIdClient(self):
        sql = "select id from info_client"
        result = self.querySelect(sql)
        massId = []
        for item in result:
            massId.append(item[0])
        return massId

    def getClientInfo(self, id_client):

        sqlInfoClient = "SELECT * from info_client where id=?"

        return self.querySelect(sqlInfoClient, (id_client,))

    def getAllInfoCity(self):
        self.comboBoxCity.clear()
        sqlInfoCity = "SELECT * from city"
        self.allInfoCity = self.querySelect(sqlInfoCity)
        self.allOnlyCity = []
        for index in range(len(self.allInfoCity)):
            self.allOnlyCity.append(self.allInfoCity[index][1])

        self.comboBoxCity.addItems(self.allOnlyCity)

    def getDateStatusForCalendar(self):
        sqlDateStatus = "SELECT date_status from work_info_client"
        self.dateStatusForCalendar = self.querySelect(sqlDateStatus)

    def getDateStatusClientForCalendar(self, id_client):
        sqlDateStatus = "SELECT date_status from work_info_client where id_client = ?"
        result = self.querySelect(sqlDateStatus, (id_client,))
        return result

    def getWorkForCalendar(self, date):
        sqlWorkInfo = "SELECT work from work_info_client where date_status = ?"
        self.workForCalendar = self.querySelect(sqlWorkInfo, date)

    def getStatusForCalendar(self):
        sqlStatusInfo = "SELECT status from work_info_client"
        self.statusForCalendar = self.querySelect(sqlStatusInfo)

    def getInfoDebts(self):
        sqlDateStatus = "select id_client, date_status, status, work from work_info_client where status = 0"
        resultDateStatus = self.querySelect(sqlDateStatus)
        dateCurrent = self.createDateObject(date.today().strftime("%d.%m.%Y"))
        massClientToForgotten = []
        for item in resultDateStatus:
            dateStatus = self.createDateObject(item[1])
            if dateStatus.daysTo(dateCurrent) > 30 and item[2] == 0 and item[3] == 0:
                massClientToForgotten.append(item[0])
        return massClientToForgotten

    #

    def insertFillInfo(self, id_client):
        date_birthday = self.dateEditBirthDay.text()
        place_residence = self.lineEditPlaceResidence.text()
        extend_work_info = self.lineEditExtendWorkInfo.text()
        approval = self.checkBoxApproval.checkState()
        contract = self.checkBoxContract.checkState()
        contract_agreement = self.checkBoxContractAgreement.checkState()
        declaration = self.checkBoxDeclaration.checkState()
        receipt = self.checkBoxReceiptOrder.checkState()
        sqlDocFillInfo = "INSERT INTO doc_fill_info(id_client, date_birthday, place_residence, extend_work_info, approval, contract, contract_agreement, declaration, receipt) VALUES (?,?,?,?,?,?,?,?,?) "
        self.queryInsert(sqlDocFillInfo, (
            id_client, date_birthday, place_residence, extend_work_info, approval, contract, contract_agreement,
            declaration,
            receipt))

    def insertAddressInfo(self, id_client):
        city = self.getIdCity(self.comboBoxCity.currentText())
        address = self.lineEditAddress.text().title()
        sqlAddressInfo = "INSERT INTO address_info_client(id_client,id_city,address) VALUES (?,?,?)"
        self.queryInsert(sqlAddressInfo, (id_client, city, address))

    def insertDocInfo(self, id_client):

        passDocSeries = self.lineEditPassDocSeries.text()
        passDocDate = self.dateEditPassDocDate.text()
        passDocInfo = self.lineEditPassDocInfo.text()
        passSnils = self.lineEditPassSnils.text()

        sqlDocInfo = "INSERT INTO doc_info_client(id_client,series_pass,date_pass,info_pass,snils) VALUES (?,?,?,?,?)"
        self.queryInsert(sqlDocInfo, (id_client, passDocSeries, passDocDate, passDocInfo, passSnils))

    def insertWorkInfo(self, id_client):

        prepayment = self.checkBoxPrepayment.checkState()
        remains = self.checkBoxRemains.checkState()
        work = 0
        dateWork = date.today().strftime("%d.%m.%Y")
        status = 0
        dateStatus = date.today().strftime("%d.%m.%Y")
        info = self.textEditInfo.toPlainText()
        sqlWorkInfo = "INSERT INTO work_info_client(id_client,prepayment,remains,work,date_work,status,date_status,info) VALUES " \
                      "(?,?,?,?,?,?,?,?) "
        self.queryInsert(sqlWorkInfo, (id_client, prepayment, remains, work, dateWork, status, dateStatus, info))

    def insertInfoClient(self):

        try:
            service = self.comboBoxProvideServices.currentIndex()
            surName = self.lineEditSurnameAdd.text()
            name = self.lineEditNameAdd.text()
            middleName = self.lineEditMiddleNameAdd.text()
            telefone = self.lineEditTelefoneAdd.text()
            self.createPath = os.path.join(os.getcwd(),
                                           self.comboBoxProvideServices.currentText(),
                                           self.comboBoxCity.currentText(),
                                           self.lineEditAddress.text() + " " + self.lineEditSurnameAdd.text())
            sqlInfoClient = "INSERT INTO info_client(sur_name,name,middle_name,telefone,path_folder,service) VALUES (?,?," \
                            "?,?,?,?) "

            self.queryInsert(sqlInfoClient, (
                surName, name, middleName, telefone, os.path.abspath(self.createPath), int(service)))
            lastIndexCleint = self.lastIndex

            self.insertAddressInfo(lastIndexCleint[0])

            self.insertWorkInfo(lastIndexCleint[0])
            self.updateDateStatusClient(date.today().strftime("%d.%m.%Y"), lastIndexCleint[0])
            self.insertDocInfo(lastIndexCleint[0])

            self.insertFillInfo(lastIndexCleint[0])

            os.makedirs(os.path.abspath(self.createPath))
            self.docCompilation()


        except BaseException as text:
            self.msgInfo("Ошибка: " + str(text))

    def insertCity(self):
        check = False
        sqlInsertCity = "INSERT INTO city(city_name) values(?)"
        dlg = QInputDialog(self)
        dlg.setInputMode(QInputDialog.TextInput)
        dlg.setWindowTitle("Добавить населенный пункт")
        dlg.setLabelText("")
        dlg.resize(300, 100)
        ok = dlg.exec_()
        city = dlg.textValue()
        if ok:
            for i in range(len(self.allOnlyCity)):
                if self.allOnlyCity[i] == city:
                    self.msgInfo("Найден похожий населенный пункт")
                    check = False
                    break
                else:
                    check = True
        if check:
            self.queryInsert(sqlInsertCity, (city,))
            self.msgInfo("Населенный пункт добавлен")
            self.comboBoxCity.clear()
            self.getAllInfoCity()

    #

    def updateTableClientToId(self, massIdClient):

        self.calendarWork()
        self.getAllClientInfo()
        self.tableWidget.setRowCount(0)
        count = 0
        massIdClient.reverse()
        for item in massIdClient:
            self.tableWidget.insertRow(self.tableWidget.rowCount())
            clientInfo = self.getClientInfo(item)
            workInfo = self.getWorkInfoClient(item)
            addressInfo = self.getAddressInfoClient(item)
            self.comboBoxStatus.setCurrentIndex(workInfo[0][6])
            self.comboBoxStatus.setCurrentIndex(workInfo[0][4])
            statusText = self.comboBoxStatus.currentText()
            statusWorkText = self.comboBoxStatus.currentText()
            nameCity = self.getNameCity(str(addressInfo[0][2]))
            addressText = addressInfo[0][3]
            if nameCity != []:
                addressCityText = nameCity[0][0]
            else:
                addressCityText = ""
            surNameText = clientInfo[0][1]
            nameText = clientInfo[0][2]
            middleNameText = clientInfo[0][3]
            telefoneText = clientInfo[0][4]
            self.comboBoxProvideServices.setCurrentIndex(clientInfo[0][6])
            serviceText = self.comboBoxProvideServices.currentText()

            item = QTableWidgetItem()
            item.setText(statusText)  # status
            self.tableWidget.setItem(count, 0, item)

            item = QTableWidgetItem()
            item.setText(statusWorkText)  # status
            self.tableWidget.setItem(count, 1, item)

            item = QTableWidgetItem()
            item.setText(addressCityText)  # city
            self.tableWidget.setItem(count, 2, item)

            item = QTableWidgetItem()
            item.setText(addressText)  # address
            self.tableWidget.setItem(count, 3, item)

            item = QTableWidgetItem()
            item.setText(surNameText)  # surName
            self.tableWidget.setItem(count, 4, item)

            item = QTableWidgetItem()
            item.setText(nameText)  # name
            self.tableWidget.setItem(count, 5, item)

            item = QTableWidgetItem()
            item.setText(middleNameText)  # middleName
            self.tableWidget.setItem(count, 6, item)

            item = QTableWidgetItem()
            item.setText(telefoneText)  # telefone
            self.tableWidget.setItem(count, 7, item)

            item = QTableWidgetItem()
            item.setText(serviceText)  # service
            self.tableWidget.setItem(count, 8, item)

            item = QTableWidgetItem()
            item.setText(str(clientInfo[0][0]))  # id
            self.tableWidget.setItem(count, 9, item)

            count = count + 1

    def updateInfoClient(self, sur_name, name, middle_name, telefone, service, id_client):
        sqlUpdateInfoClient = "UPDATE info_client SET sur_name = ?, name = ?, middle_name = ?, telefone = ?, service = ? where id = ?"
        self.queryInsert(sqlUpdateInfoClient, (sur_name, name, middle_name, telefone, service, id_client))

    def updateInfoClientPath(self, path_folder, id_client):
        sqlUpdateInfoClient = "UPDATE info_client SET path_folder = ? where id = ?"
        self.queryInsert(sqlUpdateInfoClient, (path_folder, id_client))

    def updateAddressClient(self, id_city, address, id_client):

        sqlUpdateInfoAddress = "UPDATE address_info_client SET id_city = ?, address = ? where id_client = ?"
        self.queryInsert(sqlUpdateInfoAddress, (id_city, address, id_client))

    def updateDocClient(self, series_pass, date_pass, info_pass, snils, id_client):
        sqlUpdateInfoDoc = "UPDATE doc_info_client SET series_pass = ?, date_pass = ?, info_pass = ?, snils = ? where id_client = ?"
        self.queryInsert(sqlUpdateInfoDoc, (series_pass, date_pass, info_pass, snils, id_client))

    def updateWorkClient(self, prepayment, remains, work, status, info, id_client):
        sqlUpdateInfoWork = "UPDATE work_info_client SET prepayment = ?, remains = ?, work = ?, status = ?, info = ? where id_client = ? "
        self.queryInsert(sqlUpdateInfoWork,
                         (prepayment, remains, work, status, info, id_client))

    def updateDateStatusClient(self, date_status, id_client):
        sqlUpdateInfoWork = "UPDATE work_info_client SET date_status = ? where id_client = ? "
        self.queryInsert(sqlUpdateInfoWork,
                         (date_status, id_client))

    def updateFillDocInfo(self, dateBirthDay, placeResidence, extendWorkInfo, approval, contract, contractAgreement,
                          declaration, receipt, id_client):
        sqlUpdateDocFillInfo = "UPDATE doc_fill_info SET date_birthday = ?, place_residence = ?, extend_work_info = ?, approval = ?, contract = ?, contract_agreement = ?, declaration = ?, receipt = ? where id_client = ?"
        self.queryInsert(sqlUpdateDocFillInfo, (
            dateBirthDay, placeResidence, extendWorkInfo, approval, contract, contractAgreement, declaration, receipt,
            id_client))

    def updateInfoCity(self):
        check = False
        index = ''
        sqlUpdateCity = "UPDATE city SET city_name = ? where id = ?"
        dlg = QInputDialog(self)
        dlg.setInputMode(QInputDialog.TextInput)
        dlg.setWindowTitle("Изменить населенный пункт")
        dlg.setLabelText("")
        dlg.setTextValue(self.comboBoxCity.currentText())

        dlg.resize(300, 100)
        ok = dlg.exec_()
        city = dlg.textValue()
        if ok:
            for i in range(len(self.allOnlyCity)):
                if self.allOnlyCity[i] == city:
                    self.msgInfo("Найден похожий населенный пункт")
                    check = False
                    break
                else:
                    index = self.comboBoxCity.currentIndex()
                    check = True
        if check:
            self.queryInsert(sqlUpdateCity, (city, index))
            self.msgInfo("Населенный пункт изменен")
            self.comboBoxCity.clear()
            self.getAllInfoCity()

    def updateFullInfoForClient(self):
        sur_name = self.lineEditSurname.text()
        name = self.lineEditName.text()
        middle_name = self.lineEditMiddleName.text()
        telefone = self.lineEditTelefone.text()
        service = self.comboBoxServiceChange.currentIndex()
        self.updateInfoClient(sur_name, name, middle_name, telefone, service, self.indexClient)

        series = self.lineEditPassDocSeries.text()
        datePass = self.dateEditPassDocDate.text()
        infoPass = self.lineEditPassDocInfo.text()
        snils = self.lineEditPassSnils.text()
        self.updateDocClient(series, datePass, infoPass, snils, self.indexClient)

        prepayment = self.checkBoxPrepayment.checkState()
        remains = self.checkBoxPrepayment.checkState()
        work = self.comboBoxWork.currentIndex()

        status = self.comboBoxStatus.currentIndex()

        infoClient = self.textEditInfo.toPlainText()
        self.updateWorkClient(prepayment, remains, work, status, infoClient, self.indexClient)

        dateBirthDay = self.dateEditBirthDay.text()
        placeResidence = self.lineEditPlaceResidence.text()
        extendWorkInfo = self.lineEditExtendWorkInfo.text()
        approval = self.checkBoxApproval.checkState()
        contract = self.checkBoxContract.checkState()
        contractAgreement = self.checkBoxContractAgreement.checkState()
        declaration = self.checkBoxDeclaration.checkState()
        receipt = self.checkBoxReceiptOrder.checkState()

        self.updateFillDocInfo(dateBirthDay, placeResidence, extendWorkInfo, approval, contract, contractAgreement,
                               declaration, receipt, self.indexClient)

        self.copyCompliteClient()

    def updatePathToClient(self):

        self.changeFolder()

    def updateStatusClient(self, status, id_client):
        sql = "UPDATE work_info_client SET status = ? where id_client = ?"
        self.queryInsert(sql, (status, id_client))

    def updateWorkStatus(self, work, date_work, id_client):
        sqlUpdateInfoWork = "UPDATE work_info_client SET work = ?, date_work = ? where id_client = ?"
        self.queryInsert(sqlUpdateInfoWork, (work, date_work, id_client))

    def updateFillDocStatus(self, approval, contract, contractAgreement,
                            declaration, receipt, id_client):
        sqlUpdateDocFillInfo = "UPDATE doc_fill_info SET approval = ?, contract = ?, contract_agreement = ?, declaration = ?, receipt = ? where id_client = ?"
        self.queryInsert(sqlUpdateDocFillInfo, (approval, contract, contractAgreement, declaration, receipt, id_client))

    #

    def openFolder(self, path):

        os.startfile(os.path.abspath(path))

    def copyCompliteClient(self):
        try:
            status = self.comboBoxStatus.currentIndex()
            if status == 1:
                buttonReply = QMessageBox.question(self, 'Подтверждение действия',
                                                   "Статус работ изменен: Готова; При подтверждении папка перенесется в Выполненые",
                                                   QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if buttonReply == QMessageBox.Yes:
                    path = os.path.join(os.getcwd(),
                                        self.comboBoxServiceChange.currentText(),
                                        "Выполнены",
                                        self.comboBoxCityChange.currentText(),
                                        self.lineEditAddressChange.text() + " " + self.lineEditSurname.text())
                    pathOld = self.infoClient[0][5]
                    self.updateInfoClientPath(path, self.indexClient)
                    os.makedirs(path)
                    self.pathFolderNew = path
                    for item in os.listdir(pathOld):
                        s = os.path.join(os.path.join(pathOld, item))
                        d = os.path.join(path, item)
                        if os.path.isdir(s):
                            shutil.copytree(s, d)
                        else:
                            shutil.copy2(s, d)
                    self.deleteFolder(pathOld)
                    os.startfile(path)
                    self.getAddressInfoClient(self.indexClient)
                    self.fillClientInfo()
                else:
                    self.msgInfo("Путь не изменён, найдены схожие имена")
                    self.pathFolderNew = self.infoClient[0][5]
                    self.getAddressInfoClient(self.indexClient)
        except BaseException as text:
            self.msgInfo("Ошибка: " + str(text))

    def deleteFolder(self, path):
        shutil.rmtree(path)

    def changeFolder(self):
        try:
            id_city = self.comboBoxCityChange.currentIndex()
            address = self.lineEditAddressChange.text()
            if id_city != self.infoAddressClient[0][2] or address != self.infoAddressClient[0][
                3] or self.comboBoxServiceChange.currentIndex() != self.infoClient[0][
                6] or self.comboBoxStatus.currentIndex() == 0:
                buttonReply = QMessageBox.question(self, 'Подтверждение действия', "Изменить путь к данной папке?",
                                                   QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if buttonReply == QMessageBox.Yes:
                    path = os.path.join(os.getcwd(),
                                        self.comboBoxServiceChange.currentText(),
                                        self.comboBoxCityChange.currentText(),
                                        self.lineEditAddressChange.text() + " " + self.lineEditSurname.text())
                    pathOld = self.infoClient[0][5]
                    self.updateAddressClient(id_city, address, self.indexClient)
                    self.updateInfoClientPath(path, self.indexClient)
                    self.updateStatusClient(self.comboBoxStatus.currentIndex(), self.indexClient)
                    os.makedirs(path)
                    self.pathFolderNew = path
                    for item in os.listdir(pathOld):
                        s = os.path.join(os.path.join(pathOld, item))
                        d = os.path.join(path, item)
                        if os.path.isdir(s):
                            shutil.copytree(s, d)
                        else:
                            shutil.copy2(s, d)
                    self.deleteFolder(pathOld)
                    os.startfile(path)
                    self.getAddressInfoClient(self.indexClient)
                    self.fillClientInfo()
            else:
                self.msgInfo("Путь не изменён, найдены схожие имена")
                self.pathFolderNew = self.infoClient[0][5]
                self.getAddressInfoClient(self.indexClient)
        except FileExistsError:
            self.msgInfo("Папка уже существует")
        except FileNotFoundError:
            self.msgInfo("Не найден указанный путь")

    #

    def fillClientInfo(self):
        # info data
        self.getAllInfoCity()
        self.clearChangePage()
        self.comboBoxCityChange.addItems(self.allOnlyCity)
        self.tabWidget.setCurrentIndex(2)
        row = self.tableWidget.currentRow()
        self.indexClient = self.tableWidget.item(row, 9).text()

        self.infoClient = self.getClientInfo(self.indexClient)
        self.infoPassportClient = self.getIdDocInfoClient(self.indexClient)
        self.infoWorkClient = self.getWorkInfoClient(self.indexClient)
        self.infoAddressClient = self.getAddressInfoClient(self.indexClient)
        self.infoDocFillInfo = self.getDocFillInfo(self.indexClient)

        # info city
        city = self.infoAddressClient[0][2]
        address = self.infoAddressClient[0][3]

        # info client

        surName = self.infoClient[0][1]
        name = self.infoClient[0][2]
        middleName = self.infoClient[0][3]
        telefone = self.infoClient[0][4]
        path = self.infoClient[0][5]
        service = self.infoClient[0][6]

        # passport info

        series = self.infoPassportClient[0][2]
        datePassport = self.createDateObject(self.infoPassportClient[0][3])
        infoPassport = self.infoPassportClient[0][4]
        snils = self.infoPassportClient[0][5]

        # work info

        prepaymant = self.infoWorkClient[0][2]
        paymant = self.infoWorkClient[0][3]
        workStatus = self.infoWorkClient[0][4]
        dateWorkStatus = self.createDateObject(self.infoWorkClient[0][5])
        status = self.infoWorkClient[0][6]
        dateStatus = self.createDateObject(self.infoWorkClient[0][7])
        info = self.infoWorkClient[0][8]

        # doc fill info

        dateBirthDay = self.createDateObject(self.infoDocFillInfo[0][2])
        placeResidence = self.infoDocFillInfo[0][3]
        extendWorkInfo = self.infoDocFillInfo[0][4]
        approval = int(self.infoDocFillInfo[0][5])
        contract = int(self.infoDocFillInfo[0][6])
        contractAgreement = int(self.infoDocFillInfo[0][7])
        declaration = int(self.infoDocFillInfo[0][8])
        receipt = int(self.infoDocFillInfo[0][9])

        self.comboBoxServiceChange.setCurrentIndex(int(service))
        self.comboBoxCityChange.setCurrentIndex(city)
        self.lineEditAddressChange.setText(address)
        self.lineEditPathChange.setText(path)

        self.lineEditSurname.setText(surName)
        self.lineEditName.setText(name)
        self.lineEditMiddleName.setText(middleName)
        self.lineEditTelefone.setText(telefone)

        self.lineEditPassDocSeries.setText(series)
        self.dateEditPassDocDate.setDate(datePassport)
        self.lineEditPassDocInfo.setText(infoPassport)
        self.lineEditPassSnils.setText(snils)

        self.checkBoxPrepayment.setCheckState(int(prepaymant))
        self.checkBoxRemains.setCheckState(int(paymant))
        self.comboBoxWork.setCurrentIndex(workStatus)
        self.dateEditDateWork.setDate(dateWorkStatus)
        self.comboBoxStatus.setCurrentIndex(status)
        self.dateEditReception.setDate(dateStatus)
        self.textEditInfo.setText(info)

        self.dateEditBirthDay.setDate(dateBirthDay)
        self.lineEditPlaceResidence.setText(placeResidence)
        self.lineEditExtendWorkInfo.setText(extendWorkInfo)
        self.checkBoxApproval.setCheckState(approval)
        self.checkBoxContract.setCheckState(contract)
        self.checkBoxContractAgreement.setCheckState(contractAgreement)
        self.checkBoxDeclaration.setCheckState(declaration)
        self.checkBoxReceiptOrder.setCheckState(receipt)

    def changeWorkInfo(self):
        try:
            row = self.tableWidget.currentRow()
            currentDate = date.today().strftime("%d.%m.%Y")
            self.indexClient = self.tableWidget.item(row, 9).text()
            self.updateWorkStatus(1, currentDate, self.indexClient)
            self.fillClientInfo()
        except AttributeError:
            self.msgInfo("Выберите клиента из базы")

    def addClient(self):
        self.insertInfoClient()
        self.clearAddClientPage()

    def deleteClient(self):
        buttonReply = QMessageBox.question(self, 'Подтверждение действия', "Удалить выбранную запись?",
                                           QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if buttonReply == QMessageBox.Yes:
            try:
                idClient = self.tableWidget.item(self.tableWidget.currentRow(), 9).text()
                clientPathFolder = self.getClientInfo(int(idClient))
                self.deleteFolder(clientPathFolder[0][5])
                sql = "DELETE FROM info_client WHERE id=?"
                self.queryInsert(sql, (idClient,))
                self.updateTableClientToId(self.getAllIdClient())
            except PermissionError:
                self.msgError("Удаление невозможно, закройте файлы находящиеся в папке")

    def searchClient(self):

        text = ""
        sqlCityNameId = "select id from city where city_name = ?"
        sqlCity = "select id_client from address_info_client where id_city = ?"
        sqlSurName = "select * from info_client where sur_name || telefone like ?"
        sqlAddress = "select * from address_info_client where address like ?"
        massIdClient = []
        check = False

        for city in self.allInfoCity:
            if fuzz.partial_ratio(self.lineEditSearch.text(), city[1]) > 80:
                text = city[1]
                check = True
                break

        if check:
            resultCityId = self.querySelect(sqlCityNameId, (text,))
            resultCity = self.querySelect(sqlCity, (resultCityId[0][0],))
            for idClient in resultCity:
                massIdClient.append(idClient[0])

        text = '%' + self.lineEditSearch.text() + '%'
        text = text.title()
        resultSurName = self.querySelect(sqlSurName, (text,))
        resultAddress = self.querySelect(sqlAddress, (text,))

        if len(resultSurName) > 0:
            for idClient in resultSurName:
                massIdClient.append(idClient[0])
        if len(resultAddress) > 0:
            for idClient in resultAddress:
                massIdClient.append(idClient[0])
        if len(massIdClient) > 0:
            massIdClient.reverse()
            self.updateTableClientToId(massIdClient)

    def clearAddClientPage(self):
        self.comboBoxProvideServices.setCurrentIndex(0)
        self.comboBoxCity.setCurrentIndex(0)
        self.lineEditAddress.clear()
        self.lineEditSurnameAdd.clear()
        self.lineEditNameAdd.clear()
        self.lineEditMiddleNameAdd.clear()
        self.lineEditTelefoneAdd.clear()

    def clearChangePage(self):

        self.comboBoxCityChange.clear()
        self.lineEditSurname.clear()
        self.lineEditName.clear()
        self.lineEditMiddleName.clear()
        self.lineEditTelefone.clear()

        self.lineEditPassDocSeries.clear()
        self.dateEditPassDocDate.clear()
        self.lineEditPassDocInfo.clear()
        self.lineEditPassSnils.clear()

        self.dateEditDateWork.clear()

        self.dateEditReception.clear()
        self.textEditInfo.clear()

        self.checkBoxRemains.setCheckState(0)
        self.checkBoxPrepayment.setCheckState(0)
        self.comboBoxServiceChange.setCurrentIndex(0)

        self.dateEditBirthDay.clear()
        self.lineEditPlaceResidence.clear()
        self.lineEditExtendWorkInfo.clear()

        self.checkBoxApproval.setCheckState(0)
        self.checkBoxContract.setCheckState(0)
        self.checkBoxContractAgreement.setCheckState(0)
        self.checkBoxDeclaration.setCheckState(0)
        self.checkBoxReceiptOrder.setCheckState(0)

    def clientInfoDebts(self):
        self.updateTableClientToId(self.getInfoDebts())

    def clientInfoReady(self):
        self.updateTableClientToId(self.getInfoReady())

    # calendar work

    def calendarWork(self):
        dateBuild = QDate()
        color = QColor()
        brush = QBrush()
        form = QTextCharFormat()
        self.getDateStatusForCalendar()

        for date in self.dateStatusForCalendar:

            self.getWorkForCalendar(date)

            if self.workForCalendar[0][0] == 1:

                color.setRgb(222, 245, 140)

            elif self.workForCalendar[0][0] == 0:

                color.setRgb(237, 140, 28)

            year = int(str(date[0])[6:10])
            mnt = int(str(date[0])[3:5])
            day = int(str(date[0])[0:2])
            dateBuild.setDate(year, mnt, day)
            brush.setColor(color)
            form.setBackground(brush)
            self.calendarWidget.setDateTextFormat(dateBuild, form)

    def calendarSelectWork(self):
        dateSelect = self.calendarWidget.selectedDate().toString("dd.MM.yyyy")
        massIdClient = []

        for item in self.allInfoWork:
            if item[7] == dateSelect:
                massIdClient.append(item[1])
        massIdClient.reverse()
        self.updateTableClientToId(massIdClient)

    def calendarSelectClient(self):
        row = self.tableWidget.currentRow()
        idClient = int(self.tableWidget.item(row, 9).text())
        resultIdClient = self.getDateStatusClientForCalendar(idClient)
        self.calendarWidget.setSelectedDate(self.createDateObject(resultIdClient[0][0]))

    # doc work

    def fillDocClient(self):
        try:
            getPath = self.getClientInfo(self.indexClient)
            path = getPath[0][5]
            docList = listdir(path)
            address = self.lineEditAddress.text()
            number_contact_agreement = self.indexClient
            date_contract_agreement = self.dateEditReception.text()
            sur_name = self.lineEditSurname.text()
            name = self.lineEditName.text()
            middle_name = self.lineEditMiddleName.text()
            service_extended = self.lineEditExtendWorkInfo.text()
            date_burn = self.dateEditBirthDay.text()
            place_residence = self.lineEditPlaceResidence.text()
            series_pass = self.lineEditPassDocSeries.text()
            date_pass = self.dateEditPassDocDate.text()
            info_pass = self.lineEditPassDocInfo.text()
            snils = self.lineEditPassSnils.text()
            telefone = self.lineEditTelefone.text()
            dateObject = self.createDateObject(date_contract_agreement)
            check = 0
            textErrorMsg = ""
            getFillDoc = ""
            if series_pass == "":
                textErrorMsg = "Серия/номер паспорта " + textErrorMsg
            else:
                check = check + 1
            if info_pass == "":
                textErrorMsg = "Кем выдан " + textErrorMsg
            else:
                check = check + 1
            if snils == "":
                textErrorMsg = "Снилс " + textErrorMsg
            else:
                check = check + 1
            if service_extended == "":
                textErrorMsg = "Подробное наименование работ " + textErrorMsg
            else:
                check = check + 1
            if place_residence == "":
                textErrorMsg = "Место проживания " + textErrorMsg
            else:
                check = check + 1
            if check == 5:
                for doc in docList:
                    if doc == "Согласие.docx" and self.checkBoxApproval.checkState() == 2:  # Согласие
                        docContract = DocxTemplate(os.path.join(path, "Согласие.docx"))
                        context = {
                            'number_contact_agreement': str(self.indexClient) + "-" + str(dateObject.day()) + "/" + str(
                                dateObject.month()),
                            'date_contract_agreement': date_contract_agreement,
                            'sur_name': sur_name,
                            'name': name,
                            'middle_name': middle_name,
                            'service_extended': service_extended,
                            'date_burn': date_burn,
                            'place_residence': place_residence,
                            'series_pass': series_pass,
                            'date_pass': date_pass,
                            'info_pass': info_pass,
                            'snils': snils,
                            'telefone': telefone}
                        docContract.render(context)
                        docContract.save(os.path.join(path, "Согласие(Заполнен).docx"))
                    if doc == "Акт к договору подряда.docx" and self.checkBoxContractAgreement.checkState() == 2:  # Акт к договору
                        docContract = DocxTemplate(os.path.join(path, "Акт к договору подряда.docx"))
                        context = {
                            'number_contact_agreement': str(self.indexClient) + "-" + str(dateObject.day()) + "/" + str(
                                dateObject.month()),
                            'date_contract_agreement': date_contract_agreement,
                            'sur_name': sur_name,
                            'name': name,
                            'middle_name': middle_name,
                            'service_extended': service_extended,
                            'date_burn': date_burn,
                            'place_residence': place_residence,
                            'series_pass': series_pass,
                            'date_pass': date_pass,
                            'info_pass': info_pass,
                            'snils': snils,
                            'telefone': telefone}
                        docContract.render(context)
                        docContract.save(os.path.join(path, "Акт к договору подряда(Заполнен).docx"))
                        getFillDoc = getFillDoc + "Акт к договору подряда "
                    if doc == "Договор подряда.docx" and self.checkBoxContract.checkState() == 2:  # Договор
                        docContract = DocxTemplate(os.path.join(path, "Договор подряда.docx"))
                        context = {
                            'number_contact_agreement': str(self.indexClient) + "-" + str(dateObject.day()) + "/" + str(
                                dateObject.month()),
                            'date_contract_agreement': date_contract_agreement,
                            'sur_name': sur_name,
                            'name': name,
                            'middle_name': middle_name,
                            'service_extended': service_extended,
                            'date_burn': date_burn,
                            'place_residence': place_residence,
                            'series_pass': series_pass,
                            'date_pass': date_pass,
                            'info_pass': info_pass,
                            'snils': snils,
                            'telefone': telefone}
                        docContract.render(context)
                        docContract.save(os.path.join(path, "Договор подряда(Заполнен).docx"))
                    if doc == "Декларация.xlsx" and self.checkBoxDeclaration.checkState() == 2:  # Декларация
                        if self.checkBoxDeclaration.checkState() == 2:
                            wb = openpyxl.load_workbook(os.path.join(path, "Декларация.xlsx"))
                            ws = wb.active
                            ws['Q144'] = sur_name
                            ws['Q145'] = name
                            ws['Q146'] = middle_name
                            ws['AB147'] = snils
                            ws['S148'] = series_pass
                            ws['D149'] = date_pass + " " + info_pass
                            wb.save(os.path.join(path, "Декларация(Заполнена).xlsx"))
                            wb.close()
                    if doc == "Квитанция.xlsx" and self.checkBoxReceiptOrder.checkState() == 2:  # Квитанция
                        wb = openpyxl.load_workbook(os.path.join(path, "Квитанция.xlsx"))
                        ws = wb.active
                        ws['C21'] = sur_name + " " + name + " " + middle_name
                        ws['C22'] = service_extended
                        ws['B24'] = "По адресу: " + address
                        ws['AB147'] = snils
                        ws['S148'] = series_pass
                        ws['D149'] = date_pass + " " + info_pass
                        ws['F14'] = str(self.indexClient) + "-" + str(dateObject.day()) + "/" + str(dateObject.month())
                        wb.save(os.path.join(path, "Квитанция(Заполнена).xlsx"))
                        wb.close()

                self.updateFillDocStatus(self.checkBoxApproval.checkState(),
                                         self.checkBoxContract.checkState(),
                                         self.checkBoxContractAgreement.checkState(),
                                         self.checkBoxDeclaration.checkState(),
                                         self.checkBoxReceiptOrder.checkState(),
                                         self.indexClient)

            else:
                self.msgInfo("Не заполнены следующие данные: " + textErrorMsg)
        except BaseException as text:
            self.msgInfo("Ошибка: " + str(text))
        except IndexError:
            self.msgInfo("Выберите клиента из базы")
        except PermissionError:
            self.msgInfo("Закройте заполняемые документы")

    def switchCaseDocToService(self, serviceNum):
        switch = {
            0: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx'],
            1: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx', 'координаты ЗУ.csv',
                'проект межевого плана.xlsx'],
            2: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx'],
            3: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx', 'координаты ЗУ.csv', 'МП.xlsx'],
            4: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx', 'координаты ЗУ.csv', 'МП.xlsx'],
            5: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx', 'координаты ЗУ.csv', 'МП.xlsx'],
            6: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx', 'координаты ЗУ.csv', 'МП.xlsx'],
            7: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx', 'координаты ОКС.csv', 'ТП.xlsx',
                'Декларация.xlsx', 'АКТ к договору подряда.docx'],
            8: ['Квитанция.xlsx', 'Согласие.docx', 'Договор подряда.docx', 'координаты ЗУ.csv', 'уточнение.xlsx'],
        }
        return switch.get(serviceNum)

    def docCompilation(self):

        numService = self.comboBoxProvideServices.currentIndex()

        for item in self.switchCaseDocToService(numService):
            s = os.path.join(self.pathDoc, item)
            d = os.path.join(self.createPath)
            absPathDoc = os.path.abspath(s)
            absPathClient = os.path.abspath(d)
            if os.path.isdir(absPathDoc):
                shutil.copytree(absPathDoc, absPathClient)
            else:
                shutil.copy2(absPathDoc, absPathClient)

    def remains(self):
        try:
            if self.checkBoxRemains.checkState() == 2:
                text = self.textEditInfo.toPlainText()
                self.textEditInfo.setText(text + " " + "Остаток выплачен в размере 5000 тысяч;")

            if self.checkBoxRemains.checkState() == 0:
                str = self.textEditInfo.toPlainText()
                firstIndex = str.index("Остаток выплачен")
                lastIndex = str.rindex("тысяч;") + 6
                strCross = str[firstIndex:lastIndex]
                strNew = str.replace(strCross, '')
                self.textEditInfo.setText(strNew)
        except ValueError:
            pass

    def prepayment(self):
        try:
            if self.checkBoxPrepayment.checkState() == 2:
                text = self.textEditInfo.toPlainText()
                self.textEditInfo.setText(text + " " + "Предоплата выплачена в размере 3000 тыcяч;")

            if self.checkBoxPrepayment.checkState() == 0:
                str = self.textEditInfo.toPlainText()
                firstIndex = str.index("Предоплата выплачена")
                lastIndex = str.rindex("тыcяч;") + 6
                strCross = str[firstIndex:lastIndex]
                strNew = str.replace(strCross, '')
                self.textEditInfo.setText(strNew)
        except ValueError:
            pass


def main():
    import ctypes
    myappid = 'mycompany.myproduct.subproduct.version'  # arbitrary string
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)

    app = QtWidgets.QApplication(sys.argv)
    window = startWindow()
    window.show()
    window.showMaximized()
    window.setWindowIcon(QIcon('img/main.ico'))
    app.setStyle('Fusion')
    window.updateTableClientToId(window.getAllIdClient())
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
